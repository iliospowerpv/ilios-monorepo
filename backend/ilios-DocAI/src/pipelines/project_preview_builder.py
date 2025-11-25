import logging
from pathlib import Path
from typing import Any, List

import pandas as pd

from src.pipelines.constants import AgreementType
from src.pipelines.mappings import agreement_type_to_config
from src.pipelines.poison_pills.base import PoisonPillsPipeline
from src.pipelines.post_processing import project_preview_postprocessing
from src.pipelines.term_extraction.pipeline import PipelineFactory
from src.pipelines.term_extraction.pipeline_config import PipelineConfig
from src.pipelines.term_summary.base import TermSummaryPipelineFactory


logger = logging.getLogger(__name__)


def prepare_project_preview(
    file_paths: List[str | Path],
    agreement_type: AgreementType,
    poison_pills_detection: bool = False,
) -> pd.DataFrame:
    """Prepare project preview based on chosen agreement type."""

    try:
        logger.info("Choose pipeline config based on agreement type")
        pipeline_config_class = agreement_type_to_config[agreement_type]
        pipeline_config: PipelineConfig = pipeline_config_class()
    except KeyError:
        raise ValueError(f"Unsupported agreement type: {agreement_type}")

    logger.info("Initialize pipeline from config")
    pipeline = PipelineFactory.create_pipeline(pipeline_config)
    logger.info("Build project preview")
    result = pipeline.build_project_preview(file_paths)
    logger.info("Summarize terms")
    result = TermSummaryPipelineFactory.create_pipeline(pipeline_config).run(result)
    logger.info("Postprocess project preview")
    result_postprocessed = project_preview_postprocessing(
        result, full_text=pipeline.processor.get_full_text()
    )
    logger.info("Poison pills detection generation")
    if poison_pills_detection:
        result_postprocessed = PoisonPillsPipeline().run(result_postprocessed)
    logger.info("Project preview is ready")
    return result_postprocessed


def convert_df_to_excel(
    df: pd.DataFrame,
    file_name: str = "output.xlsx",
    poison_pills_detection: bool = False,
) -> Any:
    """Convert DataFrame to CSV."""
    if poison_pills_detection:
        save_project_preview_with_poison_pills(df, file_name)
    else:
        with pd.ExcelWriter(file_name) as writer:
            df.to_excel(writer, sheet_name="Project Preview")

    return file_name


def save_project_preview_with_poison_pills(
    df: pd.DataFrame,
    file_name: str = "output.xlsx",
) -> None:
    """Save project preview with poison pills."""
    with pd.ExcelWriter(file_name) as writer:
        df.drop("Poison Pills Presented", axis=1).to_excel(
            writer, sheet_name="Project Preview"
        )
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    wb = load_workbook(file_name)
    ws = wb.active
    for i, row in enumerate(
        ws.iter_rows(min_row=2, max_row=df.shape[0] + 1, min_col=5, max_col=5)
    ):
        for cell in row:
            if df["Poison Pills Presented"][i] == "yes":
                cell.fill = PatternFill("solid", fgColor="FF0000")
    wb.save(file_name)


class ProjectPreviewBuilder:
    def __init__(
        self,
        file_paths: List[str | Path],
        agreement_type: AgreementType,
        poison_pills_detection: bool = False,
    ):
        self.agreement_type = agreement_type
        self.poison_pills_detection = poison_pills_detection
        self.file_paths = file_paths
        self.project_preview = None

    def get_project_preview(self) -> pd.DataFrame:
        """Build project preview."""
        if self.project_preview is None:
            self.project_preview = prepare_project_preview(
                self.file_paths,
                agreement_type=AgreementType(self.agreement_type),
                poison_pills_detection=self.poison_pills_detection,
            )

        return self.project_preview

    def save_project_preview_to_excel(self, file_name: str) -> Any:
        """Save project preview to an Excel file."""
        project_preview = self.get_project_preview()
        return convert_df_to_excel(
            project_preview, file_name, self.poison_pills_detection
        )

    def get_project_preview_dict(self) -> Any:
        """Get project preview in JSON format."""
        logger.info("Building project preview")
        project_preview = self.get_project_preview()
        logger.info("Building project preview dictionary")
        project_preview = project_preview.rename(
            columns={
                "Key Items": "key_item",
                "Legal Terms": "legal_term",
                "Value": "value",
                "Poison Pills": "poison_pill",
                "Poison Pills Detailed": "poison_pill_detailed",
            }
        )
        logger.info("Poison pills handling")
        if self.poison_pills_detection:
            project_preview["is_poison_pill"] = [
                1 if poison_pill_presented == "yes" else 0
                for poison_pill_presented in project_preview["Poison Pills Presented"]
            ]
            project_preview.drop(columns=["Poison Pills Presented"], inplace=True)
        else:
            project_preview["is_poison_pill"] = 0
        logger.info("Fill missing values")
        project_preview = project_preview.fillna("N/A")
        project_preview = project_preview.replace("", "N/A")
        logger.info("Parse to dictionary")
        return project_preview.to_dict(orient="records")
